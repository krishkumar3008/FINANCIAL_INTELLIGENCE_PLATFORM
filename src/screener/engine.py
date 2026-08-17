import os

import numpy as np
import openpyxl
import pandas as pd
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.database import get_db_connection

DEFAULT_CONFIG_PATH = "config/screener_config.yaml"
DEFAULT_DB_PATH = "nifty100.db"


def load_screener_config(config_path: str = DEFAULT_CONFIG_PATH) -> dict:
    """Loads screener YAML configuration file containing preset filters and composite weights."""
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Configuration file not found at {config_path}")
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def get_screener_universe(db_path: str = DEFAULT_DB_PATH) -> pd.DataFrame:
    """
    Loads complete company financial universe from SQLite database for the latest year (2024).
    Merges companies, sectors, financial_ratios, market_cap, profitandloss, and cashflow.
    Calculates derived metrics (e.g. revenue_cagr_3yr, cfo_pat_ratio, de_declining_yoy).
    """
    conn = get_db_connection(db_path)
    try:
        query = """
        SELECT 
            c.id AS company_id,
            c.company_name,
            c.roce_percentage AS roce_percentage,
            c.roe_percentage AS company_roe_percentage,
            s.broad_sector,
            s.sub_sector,
            s.market_cap_category,
            r24.return_on_equity_pct,
            r24.net_profit_margin_pct,
            r24.operating_profit_margin_pct,
            r24.debt_to_equity,
            r23.debt_to_equity AS debt_to_equity_2023,
            r24.interest_coverage,
            r24.icr_label,
            r24.asset_turnover,
            r24.free_cash_flow_cr,
            r24.capex_cr,
            r24.earnings_per_share,
            r24.book_value_per_share,
            r24.dividend_payout_ratio_pct,
            r24.total_debt_cr,
            r24.cash_from_operations_cr,
            r24.revenue_cagr_5yr,
            r24.pat_cagr_5yr,
            r24.eps_cagr_5yr,
            m24.market_cap_crore,
            m24.enterprise_value_crore,
            m24.pe_ratio,
            m24.pb_ratio,
            m24.ev_ebitda,
            m24.dividend_yield_pct,
            p24.sales,
            p24.operating_profit,
            p24.other_income,
            p24.interest,
            p24.net_profit,
            p24.sales AS sales_2024,
            p21.sales AS sales_2021
        FROM companies c
        LEFT JOIN sectors s ON c.id = s.company_id
        LEFT JOIN financial_ratios r24 ON c.id = r24.company_id AND r24.year = 2024
        LEFT JOIN financial_ratios r23 ON c.id = r23.company_id AND r23.year = 2023
        LEFT JOIN market_cap m24 ON c.id = m24.company_id AND m24.year = 2024
        LEFT JOIN profitandloss p24 ON c.id = p24.company_id AND p24.year = 2024
        LEFT JOIN profitandloss p21 ON c.id = p21.company_id AND p21.year = 2021
        """
        df = pd.read_sql(query, conn)
    finally:
        conn.close()

    # Fill fallback ROCE if missing
    df["roce_percentage"] = df["roce_percentage"].fillna(df["return_on_equity_pct"])

    # Helper flags
    df["is_financial"] = df["broad_sector"] == "Financials"

    # Revenue CAGR 3yr computation: ((sales_2024 / sales_2021) ** (1/3) - 1) * 100
    def calc_cagr_3yr(row):
        s21 = row["sales_2021"]
        s24 = row["sales_2024"]
        if pd.isna(s21) or pd.isna(s24) or s21 <= 0 or s24 <= 0:
            return np.nan
        return ((s24 / s21) ** (1 / 3) - 1) * 100

    df["revenue_cagr_3yr"] = df.apply(calc_cagr_3yr, axis=1)

    # CFO / PAT ratio computation
    def calc_cfo_pat(row):
        pat = row["net_profit"]
        cfo = row["cash_from_operations_cr"]
        if pd.isna(pat) or pd.isna(cfo) or pat == 0:
            return np.nan
        return cfo / pat

    df["cfo_pat_ratio"] = df.apply(calc_cfo_pat, axis=1)

    # FCF positive flag
    df["fcf_positive_flag"] = df["free_cash_flow_cr"].apply(
        lambda x: 100.0 if (pd.notna(x) and x > 0) else 0.0
    )

    # D/E declining YoY flag
    def is_de_declining(row):
        if row["is_financial"]:
            return True
        de24 = row["debt_to_equity"]
        de23 = row["debt_to_equity_2023"]
        if pd.isna(de24) or pd.isna(de23):
            return True
        return de24 <= de23

    df["de_declining_yoy"] = df.apply(is_de_declining, axis=1)

    return df


def winsorise_and_scale(
    series: pd.Series, p_low: float = 10.0, p_high: float = 90.0, invert: bool = False
) -> pd.Series:
    """
    Normalises metric series using P10/P90 winsorisation.
    Caps extreme values at lower and upper percentiles, then scales linearly to 0 - 100.
    If invert=True, lower original values yield higher scores (e.g. for Debt-to-Equity).
    """
    valid = series.dropna()
    if len(valid) == 0:
        return pd.Series(50.0, index=series.index)

    p10 = np.percentile(valid, p_low)
    p90 = np.percentile(valid, p_high)

    if p90 == p10:
        scaled = pd.Series(50.0, index=series.index)
    else:
        clipped = series.clip(lower=p10, upper=p90)
        scaled = (clipped - p10) / (p90 - p10) * 100.0

    if invert:
        scaled = 100.0 - scaled

    return scaled.fillna(50.0)


def calculate_composite_score(
    df: pd.DataFrame, config: dict | None = None
) -> pd.DataFrame:
    """
    Computes overall composite quality score (0-100 scale) and sector-relative composite score.
    Formula:
      35% Profitability (ROE 15% + ROCE 10% + NPM 10%)
    + 30% Cash Quality (FCF score 15% + CFO/PAT ratio 10% + FCF positive flag 5%)
    + 20% Growth (Revenue CAGR 5yr 10% + PAT CAGR 5yr 10%)
    + 15% Leverage (D/E score inverse 10% + ICR score 5%)
    """
    if config is None:
        config = load_screener_config()

    df_scored = df.copy()

    # Overall Winsorised Metrics (0 - 100 scale)
    roe_score = winsorise_and_scale(df_scored["return_on_equity_pct"])
    roce_score = winsorise_and_scale(df_scored["roce_percentage"])
    npm_score = winsorise_and_scale(df_scored["net_profit_margin_pct"])

    fcf_score = winsorise_and_scale(df_scored["free_cash_flow_cr"])
    cfo_pat_score = winsorise_and_scale(df_scored["cfo_pat_ratio"])
    fcf_pos_score = df_scored["fcf_positive_flag"]

    rev_cagr_score = winsorise_and_scale(df_scored["revenue_cagr_5yr"])
    pat_cagr_score = winsorise_and_scale(df_scored["pat_cagr_5yr"])

    de_score = winsorise_and_scale(df_scored["debt_to_equity"], invert=True)
    # For Financials, set D/E score to 100.0
    de_score = np.where(df_scored["is_financial"], 100.0, de_score)

    # For ICR, treat Debt Free or NaN with zero interest as 100
    icr_series = df_scored["interest_coverage"].copy()
    icr_series = np.where(
        (df_scored["icr_label"] == "Debt Free") | (df_scored["interest"] == 0),
        9999.0,
        icr_series,
    )
    icr_score = winsorise_and_scale(pd.Series(icr_series, index=df_scored.index))

    # Calculate overall weighted composite score
    df_scored["composite_quality_score"] = (
        0.15 * roe_score
        + 0.10 * roce_score
        + 0.10 * npm_score
        + 0.15 * fcf_score
        + 0.10 * cfo_pat_score
        + 0.05 * fcf_pos_score
        + 0.10 * rev_cagr_score
        + 0.10 * pat_cagr_score
        + 0.10 * de_score
        + 0.05 * icr_score
    ).round(2)

    # Calculate Sector-Relative Composite Score (Normalise score within each broad_sector)
    df_scored["sector_relative_composite_score"] = 50.0
    for sector, group in df_scored.groupby("broad_sector"):
        min_score = group["composite_quality_score"].min()
        max_score = group["composite_quality_score"].max()
        if max_score != min_score:
            rel_scores = (
                (group["composite_quality_score"] - min_score)
                / (max_score - min_score)
                * 100.0
            ).round(2)
            df_scored.loc[group.index, "sector_relative_composite_score"] = rel_scores
        else:
            df_scored.loc[group.index, "sector_relative_composite_score"] = 50.0

    return df_scored


def apply_screener_filter(
    df: pd.DataFrame, preset_key: str, config: dict | None = None
) -> pd.DataFrame:
    """
    Applies preset threshold filters to financial universe.
    Supports 15 filterable metrics and handles special rules:
    - D/E max filter: automatically skipped for Financials sector.
    - ICR min filter: treats Debt Free / zero interest as infinite (always passes).
    Returns DataFrame sorted by composite_quality_score descending.
    """
    if config is None:
        config = load_screener_config()

    preset = config.get("presets", {}).get(preset_key, {})
    filters = preset.get("filters", {})

    df_filtered = df.copy()

    for field, target in filters.items():
        if field == "de_declining_yoy" and target is True:
            df_filtered = df_filtered[df_filtered["de_declining_yoy"] == True]
            continue

        if field.endswith("_min"):
            col = field[:-4]
            if col == "interest_coverage":
                # ICR filter: passes if ICR >= min or icr_label == 'Debt Free' or interest == 0 or debt_to_equity == 0
                cond = (
                    (df_filtered["interest_coverage"] >= target)
                    | (df_filtered["icr_label"] == "Debt Free")
                    | (df_filtered["interest"] == 0)
                    | (df_filtered["debt_to_equity"] == 0)
                )
                df_filtered = df_filtered[cond]

            else:
                df_filtered = df_filtered[df_filtered[col] >= target]

        elif field.endswith("_max"):
            col = field[:-4]
            if col == "debt_to_equity":
                # D/E filter: automatically skip companies in Financials sector
                cond = df_filtered["is_financial"] | (
                    df_filtered["debt_to_equity"] <= target
                )
                df_filtered = df_filtered[cond]
            else:
                df_filtered = df_filtered[df_filtered[col] <= target]

    return df_filtered.sort_values(
        by="composite_quality_score", ascending=False
    ).reset_index(drop=True)


def export_screener_excel(
    output_path: str = "output/screener_output.xlsx",
    db_path: str = DEFAULT_DB_PATH,
    config_path: str = DEFAULT_CONFIG_PATH,
) -> None:
    """
    Generates output/screener_output.xlsx with 6 sheets (one per preset).
    Includes 20 KPI columns per sheet, sorted by composite_quality_score descending.
    Colour-codes cells: green fill for meeting preset threshold, red fill for failing.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    config = load_screener_config(config_path)
    df_raw = get_screener_universe(db_path)
    df_scored = calculate_composite_score(df_raw, config)

    # Columns to include in export (20 KPI columns + IDs)
    export_cols = [
        "company_id",
        "company_name",
        "broad_sector",
        "sub_sector",
        "composite_quality_score",
        "sector_relative_composite_score",
        "return_on_equity_pct",
        "roce_percentage",
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "debt_to_equity",
        "interest_coverage",
        "free_cash_flow_cr",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "eps_cagr_5yr",
        "pe_ratio",
        "pb_ratio",
        "dividend_yield_pct",
        "dividend_payout_ratio_pct",
        "market_cap_crore",
        "sales",
        "net_profit",
    ]

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Formatting styles
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F497D", end_color="1F497D", fill_type="solid"
    )

    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    green_font = Font(name="Arial", size=9, color="006100")

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(name="Arial", size=9, color="9C0006")

    regular_font = Font(name="Arial", size=9)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    preset_keys = [
        ("quality_compounder", "Quality Compounder"),
        ("value_pick", "Value Pick"),
        ("growth_accelerator", "Growth Accelerator"),
        ("dividend_champion", "Dividend Champion"),
        ("debt_free_blue_chip", "Debt-Free Blue Chip"),
        ("turnaround_watch", "Turnaround Watch"),
    ]

    for key, sheet_title in preset_keys:
        df_preset = apply_screener_filter(df_scored, key, config)
        ws = wb.create_sheet(title=sheet_title)

        # Select available export columns
        cols = [c for c in export_cols if c in df_preset.columns]
        ws.append(cols)

        # Apply header styling
        for col_num in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Get thresholds for cell color coding
        filters = config.get("presets", {}).get(key, {}).get("filters", {})

        # Write data rows
        for r_idx, row in df_preset[cols].iterrows():
            row_vals = row.tolist()
            ws.append(row_vals)
            curr_row = ws.max_row
            is_fin = row.get("broad_sector") == "Financials"

            for c_idx, col_name in enumerate(cols, start=1):
                cell = ws.cell(row=curr_row, column=c_idx)
                cell.font = regular_font
                cell.border = thin_border

                val = row[col_name]
                if pd.isna(val):
                    continue

                # Check if col_name has threshold condition
                min_key = f"{col_name}_min"
                max_key = f"{col_name}_max"

                meets_threshold = None

                if min_key in filters:
                    target = filters[min_key]
                    if col_name == "interest_coverage":
                        icr_lbl = row.get("icr_label")
                        int_val = row.get("interest")
                        if icr_lbl == "Debt Free" or int_val == 0 or val >= target:
                            meets_threshold = True
                        else:
                            meets_threshold = False
                    else:
                        meets_threshold = val >= target

                elif max_key in filters:
                    target = filters[max_key]
                    if col_name == "debt_to_equity":
                        if is_fin:
                            meets_threshold = True
                        else:
                            meets_threshold = val <= target
                    else:
                        meets_threshold = val <= target

                elif (
                    col_name == "revenue_cagr_3yr" and "revenue_cagr_3yr_min" in filters
                ):
                    meets_threshold = val >= filters["revenue_cagr_3yr_min"]

                elif col_name == "sales" and "sales_min" in filters:
                    meets_threshold = val >= filters["sales_min"]

                if meets_threshold is True:
                    cell.fill = green_fill
                    cell.font = green_font
                elif meets_threshold is False:
                    cell.fill = red_fill
                    cell.font = red_font

        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output_path)
    print(f"Successfully exported screener output to {output_path}")


if __name__ == "__main__":
    print("Testing Screener Engine...")
    cfg = load_screener_config()
    raw = get_screener_universe()
    scored = calculate_composite_score(raw, cfg)
    print(f"Loaded {len(scored)} companies.")
    export_screener_excel()
