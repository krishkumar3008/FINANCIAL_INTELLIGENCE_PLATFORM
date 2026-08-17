import os

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.database import get_db_connection

DEFAULT_DB_PATH = "nifty100.db"

PEER_METRICS = [
    ("return_on_equity_pct", "ROE", False),
    ("roce_percentage", "ROCE", False),
    ("net_profit_margin_pct", "NPM", False),
    ("debt_to_equity", "D/E", True),  # Inverse: lower is better
    ("free_cash_flow_cr", "FCF", False),
    ("pat_cagr_5yr", "PAT CAGR 5yr", False),
    ("revenue_cagr_5yr", "Revenue CAGR 5yr", False),
    ("eps_cagr_5yr", "EPS CAGR 5yr", False),
    ("interest_coverage", "Interest Coverage", False),
    ("asset_turnover", "Asset Turnover", False),
]


def load_peer_groups(db_path: str = DEFAULT_DB_PATH) -> pd.DataFrame:
    """Loads peer_groups mapping from database."""
    conn = get_db_connection(db_path)
    try:
        df = pd.read_sql(
            "SELECT peer_group_name, company_id, is_benchmark FROM peer_groups", conn
        )
    finally:
        conn.close()
    return df


def get_peer_metrics_universe(
    db_path: str = DEFAULT_DB_PATH, year: int = 2024
) -> pd.DataFrame:
    """
    Fetches raw financial metrics for companies in peer groups for specified year.
    """
    conn = get_db_connection(db_path)
    try:
        query = f"""
        SELECT 
            pg.peer_group_name,
            pg.is_benchmark,
            c.id AS company_id,
            c.company_name,
            c.roce_percentage,
            s.broad_sector,
            s.sub_sector,
            r.return_on_equity_pct,
            r.net_profit_margin_pct,
            r.operating_profit_margin_pct,
            r.debt_to_equity,
            r.interest_coverage,
            r.icr_label,
            r.asset_turnover,
            r.free_cash_flow_cr,
            r.capex_cr,
            r.earnings_per_share,
            r.book_value_per_share,
            r.dividend_payout_ratio_pct,
            r.total_debt_cr,
            r.cash_from_operations_cr,
            r.revenue_cagr_5yr,
            r.pat_cagr_5yr,
            r.eps_cagr_5yr,
            m.market_cap_crore,
            m.pe_ratio,
            m.pb_ratio,
            m.dividend_yield_pct,
            p.sales,
            p.net_profit,
            p.operating_profit,
            p.interest
        FROM peer_groups pg
        JOIN companies c ON pg.company_id = c.id
        LEFT JOIN sectors s ON c.id = s.company_id
        LEFT JOIN financial_ratios r ON c.id = r.company_id AND r.year = {year}
        LEFT JOIN market_cap m ON c.id = m.company_id AND m.year = {year}
        LEFT JOIN profitandloss p ON c.id = p.company_id AND p.year = {year}
        """
        df = pd.read_sql(query, conn)
    finally:
        conn.close()

    # Fill fallback ROCE if missing
    df["roce_percentage"] = df["roce_percentage"].fillna(df["return_on_equity_pct"])

    return df


def compute_peer_percentiles(
    db_path: str = DEFAULT_DB_PATH, year: int = 2024
) -> pd.DataFrame:
    """
    Computes SQL PERCENT_RANK for 10 metrics within each of 11 peer groups.
    For D/E ranking: inverts percentile (1 - PERCENT_RANK) so lower D/E = higher percentile rank.
    Returns tidy DataFrame formatted for peer_percentiles SQLite table.
    """
    df = get_peer_metrics_universe(db_path, year)
    records = []

    for group_name, group in df.groupby("peer_group_name"):
        n = len(group)
        for col_name, metric_label, is_inverse in PEER_METRICS:
            series = group[col_name].copy()

            # Handle ICR special values
            if col_name == "interest_coverage":
                series = np.where(
                    (group["icr_label"] == "Debt Free") | (group["interest"] == 0),
                    999999.0,
                    series,
                )
                series = pd.Series(series, index=group.index)

            # Compute rank
            if n <= 1:
                pct_ranks = pd.Series(1.0, index=group.index)
            else:
                # Rank method 'min': ties get lowest rank
                if is_inverse:
                    # Lower is better: smallest value gets highest rank
                    ranks = series.rank(method="min", ascending=False)
                else:
                    ranks = series.rank(method="min", ascending=True)
                pct_ranks = (ranks - 1.0) / (n - 1.0)

            for idx, row in group.iterrows():
                val = series.loc[idx]
                if val == 999999.0:
                    val_str = None
                else:
                    val_str = float(val) if pd.notna(val) else None

                records.append(
                    {
                        "company_id": row["company_id"],
                        "peer_group_name": group_name,
                        "metric": metric_label,
                        "value": val_str,
                        "percentile_rank": round(float(pct_ranks.loc[idx]), 4),
                        "year": year,
                    }
                )

    return pd.DataFrame(records)


def save_peer_percentiles_to_db(
    df_percentiles: pd.DataFrame, db_path: str = DEFAULT_DB_PATH
) -> None:
    """Saves computed peer percentile ranks to peer_percentiles table in SQLite."""
    conn = get_db_connection(db_path)
    try:
        conn.execute("DELETE FROM peer_percentiles WHERE year = 2024;")
        conn.commit()
        df_percentiles.to_sql("peer_percentiles", conn, if_exists="append", index=False)
        print(
            f"Successfully saved {len(df_percentiles)} percentile records to SQLite 'peer_percentiles' table."
        )
    finally:
        conn.close()


def get_company_peer_percentiles(
    company_id: str, db_path: str = DEFAULT_DB_PATH
) -> pd.DataFrame | str:
    """
    Retrieves percentile ranks for a single company.
    If company has no peer group assigned, returns string 'No peer group assigned' without error.
    """
    conn = get_db_connection(db_path)
    try:
        query = "SELECT * FROM peer_percentiles WHERE company_id = ? ORDER BY metric;"
        df = pd.read_sql(query, conn, params=(company_id.upper(),))
    finally:
        conn.close()

    if df.empty:
        return "No peer group assigned"
    return df


def export_peer_comparison_excel(
    output_path: str = "output/peer_comparison.xlsx",
    db_path: str = DEFAULT_DB_PATH,
    year: int = 2024,
) -> None:
    """
    Generates output/peer_comparison.xlsx with 11 sheets (one per peer group).
    Colour-codes percentile rank cells:
    - Green fill (>= 75th percentile)
    - Yellow fill (25th to 75th percentile)
    - Red fill (<= 25th percentile)
    Highlights benchmark company row with gold/amber background.
    Adds summary median row at the bottom.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df_raw = get_peer_metrics_universe(db_path, year)
    df_pct = compute_peer_percentiles(db_path, year)

    # Convert pct dataframe to lookup: (company_id, metric_label) -> percentile_rank
    pct_map = {}
    for _, r in df_pct.iterrows():
        pct_map[(r["company_id"], r["metric"])] = r["percentile_rank"]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F497D", end_color="1F497D", fill_type="solid"
    )

    gold_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    gold_font = Font(name="Arial", size=9, bold=True, color="7F6000")

    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    green_font = Font(name="Arial", size=9, color="006100")

    yellow_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )
    yellow_font = Font(name="Arial", size=9, color="9C6500")

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(name="Arial", size=9, color="9C0006")

    regular_font = Font(name="Arial", size=9)
    bold_font = Font(name="Arial", size=9, bold=True)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    thick_top_border = Border(
        top=Side(style="medium", color="1F497D"),
        bottom=Side(style="medium", color="1F497D"),
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
    )

    # 11 Peer Groups
    peer_groups = df_raw["peer_group_name"].unique()

    for group_name in sorted(peer_groups):
        group_df = df_raw[df_raw["peer_group_name"] == group_name].copy()
        # Sort benchmark first, then alphabetical
        group_df = group_df.sort_values(
            by=["is_benchmark", "company_id"], ascending=[False, True]
        )

        ws = wb.create_sheet(title=group_name[:31])  # Excel max sheet title 31 chars

        # Columns: company_id, company_name, is_benchmark, then pairs of (metric_val, metric_pct_rank)
        headers = ["Company ID", "Company Name", "Benchmark"]
        for _, metric_label, _ in PEER_METRICS:
            headers.append(f"{metric_label} (Val)")
            headers.append(f"{metric_label} (Rank %)")

        ws.append(headers)

        # Style header row
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        metric_vals_for_median = {m_label: [] for _, m_label, _ in PEER_METRICS}

        # Populate rows
        for _, row in group_df.iterrows():
            cid = row["company_id"]
            cname = row["company_name"]
            is_bm = row["is_benchmark"]

            row_data = [cid, cname, "Yes (Benchmark)" if is_bm == 1 else "No"]

            for col_name, metric_label, _ in PEER_METRICS:
                val = row[col_name]
                pct = pct_map.get((cid, metric_label), np.nan)

                row_data.append(val)
                row_data.append(pct)

                if pd.notna(val):
                    metric_vals_for_median[metric_label].append(val)

            ws.append(row_data)
            curr_row = ws.max_row

            # Apply row formatting
            for c_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=curr_row, column=c_idx)
                cell.font = gold_font if is_bm == 1 else regular_font
                cell.border = thin_border

                if is_bm == 1 and c_idx <= 3:
                    cell.fill = gold_fill

                # Percentile rank cell coloring
                if c_idx > 3 and (c_idx % 2 == 1):
                    rank_val = cell.value
                    if rank_val is not None and pd.notna(rank_val):
                        cell.number_format = "0.0%"
                        if rank_val >= 0.75:
                            cell.fill = green_fill
                            cell.font = green_font
                        elif rank_val >= 0.25:
                            cell.fill = yellow_fill
                            cell.font = yellow_font
                        else:
                            cell.fill = red_fill
                            cell.font = red_font
                    elif is_bm == 1:
                        cell.fill = gold_fill

        # Add Median Summary Row at Bottom
        median_row = ["MEDIAN", f"{group_name} Median", "-"]
        for _, metric_label, _ in PEER_METRICS:
            vals = metric_vals_for_median[metric_label]
            med_val = float(np.median(vals)) if len(vals) > 0 else np.nan
            median_row.append(med_val)
            median_row.append("-")

        ws.append(median_row)
        summary_row_idx = ws.max_row

        for c_idx in range(1, len(median_row) + 1):
            cell = ws.cell(row=summary_row_idx, column=c_idx)
            cell.font = bold_font
            cell.border = thick_top_border

        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(output_path)
    print(f"Successfully exported peer comparison report to {output_path}")


if __name__ == "__main__":
    print("Testing Peer Engine...")
    df_ranks = compute_peer_percentiles()
    save_peer_percentiles_to_db(df_ranks)
    export_peer_comparison_excel()
